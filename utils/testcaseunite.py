# -*- coding:utf-8 -*-
import os, stat
import copy
import re

import shutil
import zipfile
from os.path import join, getsize

from flask import current_app, session
from openpyxl import Workbook, load_workbook
from robot.api import TestData
from robot.parsing.model import Step

from utils.file import get_projectdirfromkey ,remove_dir
from utils.mylogger import getlogger


log = getlogger("TestCaseUnite")

def getCaseContent(cpath, cname):
    '''反写：自动化结果反写中，取得测试用例内容 '''
    if not os.path.exists(cpath):
        return "Can not find case file:"+cpath

    content = ''
    suite = TestData(source=cpath)
    for t in suite.testcase_table.tests:
        if t.name == cname:
            isHand = False
            if t.tags.value and 'Hand' in t.tags.value:
                isHand = True
            for s in t.steps:
                ststr = (' ' * 4).join(s.as_list())
                if ststr.strip() == 'No Operation':
                    continue
                if isHand:
                    if ststr.strip().startswith('#*'):
                        ststr = ststr.replace('#*', '')

                content += ststr + '\n'
    return content

def export_casezip(key, exp_filedir=''):

    dir = exp_filedir
    if dir == '':
        dir = get_projectdirfromkey(key) + '/runtime'

    zip_name = os.path.basename(key) + '.zip'
    zip_path = os.path.join(dir,zip_name)

    try:
        z = zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED)

        for dirpath, dirnames, filenames in os.walk(key):
            fpath = dirpath.replace(key, '')
            fpath = fpath and fpath + os.sep or ''
            for filename in filenames:
                z.write(os.path.join(dirpath, filename), fpath + filename)
        z.close()
    except Exception as e:
        log.error("download zip case Execpiton:{}".format(e))
        return (False, "{}".format(e))

    return (True, zip_path)

def export_casexlsx(key, db, exp_filedir=''):

    export_dir = key
    if not os.path.isdir(export_dir):
        log.error("Do not support export cases of a file:"+export_dir)
        return (False,"Do not support export cases of a file:"+export_dir )

    basename = os.path.basename(export_dir)

    dir = exp_filedir
    if dir == '':
        dir = get_projectdirfromkey(export_dir) + '/runtime'

    os.mkdir(dir) if not os.path.exists(dir) else None

    export_file = os.path.join(dir, basename+'.xlsx')

    db.refresh_caseinfo(export_dir, "Force")

    cases = []
    sql = "SELECT info_key,info_name,info_doc,info_tags FROM testcase WHERE info_key like '{}%' ;".format(key)
    res = db.runsql(sql)
    for i in res:
        (info_key, info_name,info_doc,info_tag) = i
        cases.append([info_key,info_name,info_doc,info_tag])

    wb = Workbook()

    ws = wb.active
    ws.append(["导出&导入用例："])
    ws.append(["'_'用在文件名前后，表示用例文件：如 '_用例文件名_' 表示'用例文件名.robot'"])
    ws.append(["'-'用来连接目录，没有此符号表示没有子目录：如 '目录1-目录11' 表示 '目录1/目录11'"])
    ws.append(["每个sheet的第一列，是后面用例所在的用例文件名（.robot）"])
    ws.append(["... ..."])
    ws.append(["注意：通过xlsx文件导入用例，如果是自动化用例，且用例已经存在，则只更新doc和tag，不更新用例内容"])
    ws.append(["... ..."])
    ws.append(["此'sheet'页面，不会被导入"])
    ws.append(["... ..."])
    ws.append(["Export&Import Cases："])
    ws.append(["'_'after the file name，Means a suite：'_SuiteName_' means 'SuiteName.robot'"])
    ws.append(["'-'concat the dirs，no this sign no subdir：'dir1-dir11' means 'dir1/dir11'"])
    ws.append(["First Column of each sheet，is the suite name of the case in this line（.robot）"])
    ws.append(["... ..."])
    ws.append(["Caution：Import cases from xlsx file，if it is Auto-case and it exists，then update doc and tag Only，Do not update Case content."])
    ws.append(["... ..."])
    ws.append(["This 'sheet' ，Wont be imported."])

    for c in cases:
        if not os.path.exists(c[0]):
            continue

        casecontent = getCaseContent(c[0], c[1])
        suitename = os.path.basename(c[0])

        tags = c[3].split(',')
        tags.remove('${EMPTY}') if '${EMPTY}' in tags else None

        category = "Auto"

        if "HAND" in tags or "Hand" in tags or 'hand' in tags:
            category = "Hand"
            casecontent = casecontent.replace(' '*4 + '#','')

        sheetname = _get_ws(export_dir, c[0])

        #print("Get sheete name :"+sheetname)

        #print(suitename,c[1],c[2],casecontent,c[3],category)

        if not sheetname in wb.sheetnames:
            ws = wb.create_sheet(sheetname)
            #ws = wb.active
            ws.append(["Suite_Name","Case_Name","Case_Doc","Case_Content","Case_Tag","Case_Type"])
        else:
            ws = wb[sheetname]
            #ws = wb.active
        ws.append([suitename,c[1],c[2],casecontent,c[3],category])

    os.remove(export_file) if os.path.exists(export_file) else None
    wb.save(export_file)
    log.info("Generate TestCasefile of {} to {}".format(export_dir,export_file))

    return (True, export_file)

def _get_ws(export_dir,suite_key):
    """
    return worksheet name
    suite_key= /xxx/project/TestCase/v50/1dir1/test1.robot
    expor_dir= /xxx/project/TestCase
    """

    suite_name = os.path.basename(suite_key)   #test1.robot
    suite_dir = os.path.dirname(suite_key)     #/xxx/project/TestCase/v50/1dir1

    subdir = suite_dir.split(export_dir)[1]    #/v50/1dir1
    subdir = subdir.replace('/', '-')           #_v50_1dir1
    subdir = subdir[1:]                        #v50_1dir1

    if subdir == '':
        singal_suite = suite_name.split(".")[0]
        return "_"+singal_suite+"_"

    return subdir

def do_importfromzip(temp_file, path):

    zip_file = temp_file

    try:
        if not os.path.exists(zip_file):
            return ('fail', 'Can not find xlsx file :{}'.format(zip_file))
        if not os.path.isdir(path):
            return ('fail', 'The Node is NOT A DIR :{}'.format(path))

        app = current_app._get_current_object()

        if not zipfile.is_zipfile(zip_file):
            return ('fail', 'The file is not a zip file :{}'.format(os.path.basename(zip_file)))

        remove_dir(path) if os.path.exists(path) else None
        os.mkdir(path)

        fz = zipfile.ZipFile(zip_file, 'r')
        for file in fz.namelist():
            fz.extract(file, path)

        return ('success', path)
    except Exception as e:
        log.error("import from zip Exception:{}".format(e))
        return ("fail", "Exception occured .")

def do_importfromxlsx(temp_file, path):

    xls_file = temp_file
    dest_dir = path

    if not os.path.isdir(dest_dir):
        return ('fail','The Node is NOT A DIR :{}'.format(dest_dir))
    if not os.path.exists(xls_file):
        return ('fail', 'Can not find xlsx file :{}'.format(xls_file))
    xls_name = os.path.basename(xls_file).split('.')[0]
    dir_name = os.path.basename(dest_dir)
    if not xls_name == dir_name:
        return ('fail', 'Filename {} is not equal to dir name :{}'.format(xls_name,dest_dir))

    try:
        wb = load_workbook(xls_file)

        update_cases = 0
        unupdate_case = 0
        failedlist = []
        for stn in wb.sheetnames[1:]:
            ws = wb[stn]
            if not ws['A1'] != 'Suite_Name':
                return ('fail', 'sheet:{} A1:{} Expect:Suite_Name'.format(stn,ws['A1']))
            if not ws['B1'] != 'Case_Name':
                return ('fail', 'sheet:{} B1:{} Expect:Case_Name'.format(stn,ws['B1']))
            if not ws['C1'] != 'Case_Doc':
                return ('fail', 'sheet:{} C1:{} Expect:Case_Doc'.format(stn,ws['C1']))
            if not ws['D1'] != 'Case_Content':
                return ('fail', 'sheet:{} C1:{} Expect:Case_Content'.format(stn,ws['D1']))
            if not ws['E1'] != 'Case_Tag':
                return ('fail', 'sheet:{} C1:{} Expect:Case_Tag'.format(stn,ws['E1']))
            if not ws['F1'] != 'Case_Type':
                return ('fail', 'sheet:{} C1:{} Expect:Case_Type'.format(stn,ws['F1']))

            for r in ws.rows:
                (a,b,c,d,e,f) = r
                if a.value == 'Suite_Name':    # omit the 1st line
                    continue
                fields = [a.value if a.value else '',
                          b.value if b.value else '',
                          c.value if c.value else '',
                          d.value if d.value else '',
                          e.value if e.value else '',
                          f.value if f.value else ''
                          ]
                (done , msg) = _update_onecase(dest_dir,stn,fields)
                if done:
                    update_cases += 1
                else:
                    unupdate_case += 1
                    failedlist.append("sheet:{} suite:{} case:{} ->{}".format(stn,a.value,b.value,msg))

        return ('success', 'S:{},F:{},Failist:{}'.format(update_cases,unupdate_case,'\n'.join(failedlist)))

    except Exception as e:
        log.error("do_uploadcase Exception:{}".format(e))
        return ('fail','Deal with xlsx file fail :{}'.format(xls_file))

def _update_onecase(dest_dir,sheetname,fields):

    stn = sheetname

    DONE = False

    robotname = fields[0].split('.')[0]
    if stn.startswith('_') and stn.endswith('_'):
        if not '_'+robotname+'_' == stn:
            return (False,"Sheetname Should be same as the First column（no metter ext）:{} vs {}".format(stn,robotname) )
        robotfile = os.path.join(dest_dir, robotname+'.robot')
    else:
        subdir = stn.replace('-','/')
        robotfile = os.path.join(dest_dir, subdir, robotname+'.robot')

    file_dir = os.path.dirname(robotfile)
    os.makedirs(file_dir,exist_ok=True)

    log.info("Updating robotfile:{} with args:{}".format(robotfile,fields))

    isHand = False
    if fields[5] == '手工' or fields[5] == 'HAND' or fields[5] == 'Hand' or fields[5] == 'hand':
        isHand = True

    brandnew = "*** Settings ***\n" + \
               "*** Variables ***\n" + \
               "*** Test Cases ***\n" + \
               "NewTestCase\n"        + \
               "    [Documentation]  This is Doc \n" + \
               "    [Tags]   tag1  tag2\n" + \
               "    Log  This is a Brandnew case.\n"

    name = fields[1].strip()
    doc  = fields[2].strip()
    content = fields[3].strip()
    tags = fields[4].strip().replace('，', ',').split(',')  # Chinese characters
    if isHand:
        tags.append('Hand')
    tags = list(set(tags))

    space_splitter = re.compile(u'[ \t\xa0]{2,}|\t+')  # robot spliter

    try:
        # 如果文件不存在，直接创建文件和用例
        if not os.path.exists(robotfile):

            log.info("Case file is Not exists CREATE it :"+robotfile)
            with open(robotfile,'w') as f:
                f.write(brandnew)

            suite = TestData(source=robotfile)
            t = suite.testcase_table.tests[0]
            t.name = name
            t.tags.value = tags
            t.doc.value = doc.replace('\n','\\n')

            steps = []
            if isHand:
                lines = content.split('\n')
                for l in lines:
                    step = Step([],comment="#*"+l.strip())
                    steps.append(step)
                steps.append(Step(["No Operation"]))
            else:
                lines = content.split('\n')
                for l in lines:
                    step = Step(space_splitter.split(l.strip()))
                    steps.append(step)

            t.steps = steps  #如果用例不存在，则所有内容都更新:New test Case , Update all.

            suite.save(txt_separating_spaces=4)

            DONE = True

            return (DONE, robotfile)

        # 如果文件存在： 1 用例存在 ，2 用例不存在 :If file exits: 1 case exists ,2 Case doesnt exists.
        suite = TestData(source=robotfile)

        for t in suite.testcase_table.tests:
            if t.name == name:      # 2用例存在:Case exists
                log.info("Case file exits, case exists update it: "+name)
                t.tags.value = tags
                t.doc.value = doc.replace('\n', '\\n')

                if isHand:        #只有手工用例更新 用例内容，存在都自动化用例不更新内容
                    steps = []
                    lines = content.split('\n')
                    for l in lines:
                        step = Step([], comment="#*" + l.strip())
                        steps.append(step)
                    steps.append(Step(["No Operation"]))
                    t.steps = steps
                DONE = True
                break

        if DONE:
            suite.save(txt_separating_spaces=4)
            return (DONE,robotfile)

        # 1用例不存在， 需要新增用例: Case doesnt exists， Add new one
        suite = TestData(source=robotfile)

        if len( suite.testcase_table.tests ) > 0:
            log.info("Case file exists and Not null ,copy and modify a case as a new one.")
            t = copy.deepcopy(suite.testcase_table.tests[-1])
            t.name = name
            t.tags.value = tags
            t.doc.value = doc.replace('\n', '\\n')

            steps = []
            if isHand:
                lines = content.split('\n')
                for l in lines:
                    step = Step([], comment="#*" + l.strip())
                    steps.append(step)
                steps.append(Step(["No Operation"]))
            else:
                lines = content.split('\n')
                for l in lines:
                    step = Step(space_splitter.split(l.strip()))
                    steps.append(step)

            t.steps = steps

            suite.testcase_table.tests.append(t)

            suite.save(txt_separating_spaces=4)

            DONE = True

            return (DONE, robotfile)

        else:
            #用例文件存在，但是用例文件里面没有用例，异常
            log.warning("Case file exist but There isnt any case. Remove it and create a new one.")
            os.remove(robotfile)
            with open(robotfile,'w') as f:
                f.write(brandnew)

            suite = TestData(source=robotfile)
            t = suite.testcase_table.tests[0]
            t.name = name
            t.tags.value = tags
            t.doc.value = doc.replace('\n','\\n')

            steps = []
            if isHand:
                lines = content.split('\n')
                for l in lines:
                    step = Step([],comment="#*"+l.strip())
                    steps.append(step)
                steps.append(Step(["No Operation"]))
            else:
                lines = content.split('\n')
                for l in lines:
                    step = Step(space_splitter.split(l.strip()))
                    steps.append(step)

            t.steps = steps

            suite.save(txt_separating_spaces=4)

            DONE = True

            return (DONE, robotfile)

    except Exception as e:
        log.error("updateOneCase Execption:{}".format(e))
        return(DONE,"ErrorOccur:{}".format(e))



if __name__ == '__main__':

    #from utils.dbclass import TestDB
    #myDB = TestDB('/Users/tester/PycharmProjects/uniRobotDev/.beats')

    #print(export_cases("/Users/tester/PycharmProjects/uniRobotDev/.beats/workspace/tbdsadmin/RobotTbds/TestCase/v41",myDB))
    #print("sss" + _get_ws("/Users/tester/PycharmProjects/uniRobotDev/.beats/workspace/Admin/Demo_Project/RobotTestDemo/TestCase/903dir1","/Users/tester/PycharmProjects/uniRobotDev/.beats/workspace/Admin/Demo_Project/RobotTestDemo/TestCase/903dir1/test1.robot")+"xxx")
    print( do_importfromxlsx('/Users/tester/PycharmProjects/uniRobotDev/work/runtime/testloadcase.xlsx','/Users/tester/PycharmProjects/uniRobotDev/work/workspace/Admin/123TestLoad/testloadcase'))
